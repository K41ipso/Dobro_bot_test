import openpyxl

wookbook = openpyxl.load_workbook(r'phrases.xlsx')

worksheet = wookbook.active

phrases_list = []

for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        phrases_list.append(col[i].value)
print(phrases_list)